import numpy as np
from scipy import signal
from scipy.signal import find_peaks
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import os
import openpyxl 
from scipy.integrate import simps
import adi
import pandas as pd
from tqdm import tqdm
from scipy.stats import mannwhitneyu

font_path = 'C:\\Windows\\Fonts\\simsun.ttc'  # 宋體
font_prop = FontProperties(fname=font_path)

mode = 'show'   # 'show' or 'save'
patient = 'patient' # 'normal' or 'patient'


def butter(DataL, DataR, cut_low, cut_high, sample_rate, order):
    nyqs = sample_rate * 0.5
    H_cut = cut_high / nyqs
    L_cut = cut_low / nyqs
    sos = signal.butter(order, [L_cut, H_cut], analog=False, btype='bandpass', output='sos')
    Filter_Left = signal.sosfiltfilt(sos, DataL)
    Filter_Right = signal.sosfiltfilt(sos, DataR)
    return Filter_Left, Filter_Right

def bassel(DataL, DataR, cut_low, cut_high, sample_rate, order):
    #貝塞爾
    nyqs = sample_rate * 0.5
    H_cut = cut_high / nyqs
    L_cut = cut_low / nyqs
    sos=signal.bessel(order, [L_cut, H_cut] ,  btype='bandpass',  analog=False,  output='sos')
    Filter_Left = signal.sosfiltfilt(sos,  DataL) 
    Filter_Right = signal.sosfiltfilt(sos,  DataR) 

    return Filter_Left, Filter_Right

def find_peak(Filter_Data):
    valley_x, valley_y = find_peaks(Filter_Data * -1, height=0, distance=500)
    cardiac_cycle = np.diff(valley_x)
    peaks_x, peaks_y = find_peaks(Filter_Data, height=0, distance=500)
    peaks_y = peaks_y['peak_heights']
    valley_y = valley_y['peak_heights']
    return peaks_x, peaks_y, valley_x, valley_y, cardiac_cycle

def derivative(Data, Level, values=[]):
    result = np.gradient(Data)

    if Level == 0:
        values.append(Data)
        return 0
    else:
        values.append(Data)
        return derivative(result, Level - 1, values)

def process_wave(cycle):
    values = []
    derivative(cycle, 3, values)
    values = np.array(values)
    origin, derivative_1, derivative_2, derivative_3 = values
    derivative_1 = derivative_1 * 50
    derivative_2 = derivative_2 * 5000
    derivative_3 = derivative_3 * 100000

    return [derivative_1, derivative_2, derivative_3]

def Find_Path(path):

    File_path = []

    #find all Data_file path 
    for root,  subfolders,  filenames in os.walk(path):
        for filename in filenames:
            if filename.endswith('.txt'):
                continue
            filepath = root +'/'+ filename
            new_filepath = filepath.replace("\\", "/")
            File_path.append(new_filepath)

    return File_path

def get_Imformation(path,locate, imformation=[]):
    test=path.split('/')
    Name = test[locate[0]]
    Date = test[locate[1]]
    State_check = test[locate[2]]
    if State_check =='易堵':
        State = '0'
    else: 
        State = '1'

    file_name = test[len(test)-1]
    name_check = file_name.find('R')
    if name_check != -1:
        Status = 'Right'
    else:
        Status = 'Left'
    
    imformation =[Name, Date, State, Status]

    return imformation, Name

def plot(cycle_1, cycle_2, parameter, Name,i):
    p = 0
    def on_key(event):
        if event.key == 'z':
            plt.close()
    if p == 0:
        plt.plot(cycle_1, label='Left PPG')
        plt.plot(cycle_2, label='Right PPG')
    if p == 1:
        print('not yet')

    plt.title(f'{Name}, {i + 1}th Left_Right',fontproperties=font_prop)
    plt.legend()
    plt.grid()
    if mode == 'show':
        plt.show()
    else:
        plt.savefig(f'F:\\Python\\PPG\\Cycle\\{Name}, {i + 1}th Left_Right.jpg')

    plt.gcf().canvas.mpl_connect('key_press_event', on_key)
    plt.close()

def calculate_d1(cycle, Name, i):
    derivative = process_wave(cycle)
    # d1_peak,_ = find_peaks(derivative[0], height=0, distance=800)
    # d1_valley,_ = find_peaks(derivative[0] * -1, height=0, distance=800)
    x = np.linspace(0, len(cycle), len(cycle))
    plt.plot(cycle)
    plt.plot(derivative[0])
    plt.plot(derivative[1])

    #calculate_d2(cycle, derivative, Name, i) 
    find_fdppg_features(derivative[0],x)

def find_fdppg_features(fdppg,x, threshold_max=0.5, threshold_min=0.5):
    """
    找出FDPPG信號中的特徵點。
    
    參數：
    - fdppg: 一階導數的PPG信號
    - threshold_max: 用於識別最大峰值的閾值（通常設置為最大值的50%）
    - threshold_min: 用於識別最小峰值的閾值（通常設置為最小值的80%）
    
    返回：
    - max_peaks: 最大峰值的索引列表
    - min_peaks: 最小峰值的索引列表
    - local_extreme_points: 局部極值點的索引列表
    """
    max_peaks = []
    min_peaks = []
    local_extreme_points = []

    max_threshold = np.max(fdppg) * threshold_max
    min_threshold = np.min(fdppg) * threshold_min
    
    # 找出最大峰值
    max_peaks = find_peaks(fdppg, height=0, distance=500)[0]
    
    # 找出最小峰值
    min_peaks = find_peaks(fdppg * -1, height=(0, max_peaks[0]), distance=500)[0]
    
    # 找出局部極值點（通常在最小峰值和下一個零交叉點之間）
    # for i in range(len(min_peaks) - 1):
    #     region = fdppg[min_peaks[i]:min_peaks[i + 1]]
    #     local_extreme = min_peaks[i] + np.argmin(region)
    #     local_extreme_points.append(local_extreme)
    
    print(max_peaks, min_peaks, local_extreme_points)
    plt.plot(x[max_peaks], fdppg[max_peaks], '*', label='max_peaks')
    plt.plot(x[min_peaks], fdppg[min_peaks], '*', label='min_peaks')
    plt.plot(x[local_extreme_points], fdppg[local_extreme_points], '*', label='local_extreme_points')
    plt.show()
    #return max_peaks, min_peaks, local_extreme_points

def calculate_d2(cycle, derivative, Name, i):
    TDPPG_x = np.where(np.diff(np.sign(derivative[2])))[0]
    # zero_TDPPG,_ = find_peaks(derivative[1] * -1, height=0)
    TDPPG_y = derivative[2][TDPPG_x]
    x = np.linspace(0, len(derivative[2]), len(derivative[2]))
    
    plt.plot(derivative[1])
    plt.plot(x[TDPPG_x], derivative[1][TDPPG_x], '*', label='second Derivative')
    plt.title(f'{Name}, {i + 1}th Left_Right',fontproperties=font_prop)
    plt.legend()
    plt.grid()
    # if mode == 'show':
    #     #plt.show()
    # else:
    #     plt.savefig(f'F:\\Python\\PPG\\Cycle\\{Name}, {i + 1}th Left_Right.jpg')
    # plt.close()
    find_sdppg_features(derivative[1])

def find_sdppg_features(sdppg, threshold_a=0.45):
    """
    找出SDPPG信號中的特徵點。
    
    參數：
    - sdppg: 二階導數的PPG信號
    - threshold_a: 用於識別'a'峰值的閾值（通常設置為最大值的45%）
    
    返回：
    - a_peaks: 'a'波的索引列表
    - b_peaks: 'b'波的索引列表
    - c_peaks: 'c'波的索引列表
    - d_peaks: 'd'波的索引列表
    - e_peaks: 'e'波的索引列表
    """
    a_peaks = []
    b_peaks = []
    c_peaks = []
    d_peaks = []
    e_peaks = []

    a_threshold = np.max(sdppg) * threshold_a
    
    # 找出'a'波峰值
    for i in range(1, len(sdppg) - 1):
        if sdppg[i] > a_threshold and sdppg[i] > sdppg[i - 1] and sdppg[i] > sdppg[i + 1]:
            a_peaks.append(i)
    
    # 找出'b'波峰值
    for i in range(len(a_peaks) - 1):
        region = sdppg[a_peaks[i]:a_peaks[i + 1]]
        b_peak = a_peaks[i] + np.argmin(region)
        b_peaks.append(b_peak)
    
    # 'c'波和'd'波之間有時可能重疊，需要特別處理
    for i in range(len(b_peaks) - 1):
        region = sdppg[b_peaks[i]:b_peaks[i + 1]]
        c_peak = b_peaks[i] + np.argmax(region)
        c_peaks.append(c_peak)
        
        region = sdppg[c_peak:b_peaks[i + 1]]
        d_peak = c_peak + np.argmax(region)
        d_peaks.append(d_peak)
    
    # 找出'e'波峰值
    for i in range(len(b_peaks) - 1):
        region = sdppg[b_peaks[i]:b_peaks[i + 1]]
        e_peak = b_peaks[i] + np.argmin(region)
        e_peaks.append(e_peak)
    
    print(a_peaks, b_peaks, c_peaks, d_peaks, e_peaks)
    plt.show()

def calculate_cycle(cycle,cycle_cut, peak):
    B1 = max(peak[1])
    B2 = 0
    B3 = len(cycle_cut[0])
    B6 = peak[0][1] - peak[0][0]
    B7 = len(cycle_cut[0]) - peak[0][0]
    B10 = peak[0][0]
    
    imformation = [B1, B2, B3, B6, B7, B10]
    return imformation

def Write_Excel(All_imformation):
    workbook = openpyxl.load_workbook("F:\\Python\\PPG\\output.xlsx")
    sheet1 = workbook.worksheets[0]

    Data_Row = sheet1.max_row+1

    length = len(All_imformation)

    for i in range(1, length+1):
        sheet1.cell(Data_Row , i).value= All_imformation[i-1]
    workbook.save("F:\\Python\\PPG\\output.xlsx")

def main():
    channel1_id = 2
    channel2_id = 4
    record_id = 1

    if patient == 'normal':
        File_path = Find_Path("F:\\正常人Data") #!正常人
    else:
        File_path = Find_Path("F:\\第二次收案") #!病患
    print('找到資料筆數', len(File_path))

    Features = []
    df_c = pd.DataFrame()
    for j, path in tqdm(enumerate(File_path), total=len(File_path), desc='Processing'):
        Data = adi.read_file(path)

        Right = Data.channels[channel1_id - 1].get_data(record_id)
        Left = Data.channels[channel2_id - 1].get_data(record_id)

        Filter_Left,Filter_Right = butter(Left, Right, 0.5, 9, 1000, 4)

        L_wave = Filter_Left[100000:120000] * 10
        R_wave = Filter_Right[100000:120000] * 10

        L_valley_x, L_valley_y = find_peaks(L_wave * -1, height=0, distance=150)
        R_valley_x, R_valley_y = find_peaks(R_wave * -1, height=0, distance=150)

        L_valley_y = L_valley_y['peak_heights']
        R_valley_y = R_valley_y['peak_heights']

        if patient == 'normal':
            Imformation,Name = get_Imformation(path,locate=[2,3,4]) #!正常人
        else:
            Imformation,Name = get_Imformation(path,locate=[3,4,2]) #!病患

        if len(L_valley_x) > len(R_valley_x): #找最小的cycle
            min_cycle = len(R_valley_x)
        else:
            min_cycle = len(L_valley_x)
        
        for i in range(0,min_cycle - 2,2):
            diff = np.abs(L_valley_x[i] - R_valley_x[i]) #time diff
            L_cycle = L_wave[L_valley_x[i]:L_valley_x[i + 2]] #two cycle
            L_cycle_cut = [L_wave[L_valley_x[i]:L_valley_x[i + 1]], L_wave[L_valley_x[i + 1]:L_valley_x[i + 2]]] #divide 2

            L_peaks_x, L_peaks_y = find_peaks(L_cycle, height=0, distance=500)
            L_peaks_y = L_peaks_y['peak_heights']
            L_peak = [L_peaks_x, L_peaks_y]

            R_cycle = R_wave[L_valley_x[i]:L_valley_x[i + 2]]
            R_cycle_cut = [R_wave[R_valley_x[i]:R_valley_x[i + 1]], R_wave[R_valley_x[i + 1]:R_valley_x[i + 2]]]  # vivide 2

            R_peaks_x, R_peaks_y = find_peaks(R_cycle, height=0, distance=500)
            R_peaks_y = R_peaks_y['peak_heights']
            R_peak = [R_peaks_x, R_peaks_y]

            if len(L_cycle) < 1100 or len(L_peaks_y) != 2 or len(R_peaks_y) != 2 or len(R_cycle) < 1100:
                continue

            if L_peaks_y[0] < 0.5:
                L_cycle *= 0.5 / L_peaks_y[0]
                L_peaks_y[0] = 0.5
                L_peaks_y[1] = 0.5

            if R_peaks_y[0] < 0.5:
                R_cycle *= 0.5 / R_peaks_y[0]
                R_peaks_y[0] = 0.5
                R_peaks_y[1] = 0.5
            
            parameter = calculate_cycle(L_cycle, L_cycle_cut, L_peak)
            
            
            calculate_d1(L_cycle_cut[0], Name, i)
            
            min_len = min(len(L_cycle), len(R_cycle))
            u_statistic, p = mannwhitneyu(L_cycle_cut[0][0:min_len], R_cycle_cut[0][0:min_len])
            #print(u_statistic, p)
            parameter.append(p)
            
            #All_imformation = Imformation + parameter
            #Write_Excel(All_imformation)
            #plot(L_cycle, R_cycle, 0, Name,i)



if __name__ == '__main__':
    main()
